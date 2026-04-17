"""
make_pivot_table.py
====================================
Generates three pivot tables and a heatmap from a World Bank livestock projects CSV.

Outputs (saved to the same directory as the input file):
  - livestock_pivot_tables.xlsx  (3 sheets: Interventions Pivot, Outcomes Pivot, Co-occurrence Table)
  - intervention_outcome_heatmap.png

Usage:
  1. Set INPUT_CSV_PATH below to the path of your CSV file
  2. Run: python make_pivot_table.py
"""

import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt



# !! IMPORTANT: SET YOUR FILE PATH HERE !!
INPUT_CSV_PATH = "/Users/bipu/Desktop/AIMForScale/World Bank Documents PAD/PID/Pivot Tables New/world_bank_livestock_projects.csv"


# ──────────────────────────────────────────────
# 1. CATEGORY DEFINITIONS
# ──────────────────────────────────────────────
INTERVENTION_CATS = [
    'Crop Management, Technology, R&D',
    'Livestock Management & Technologies',
    'Forestry Management & Technologies',
    'Aquaculture Management & Technologies',
    'Ecosystem services',
    'Advisory and extension',
    'Markets & finance',
    'Food processing & storage',
    'Behavior communication change & social programs',
    'Policy & Regulatory',
    'Public health',
    'Rural/public infrastructure',
]

OUTCOME_CATS = [
    'Sustainable Economic Growth',
    'Resilience',
    "Inclusivity & Women's Empowerment",
    'Environmental Sustainability',
    'Food Security & Nutrition',
    'Inducing Policy Change',
]

# Alias dictionaries map all raw variants (lowercased) → canonical name.
# Add new variants here if your data changes.
INT_ALIASES = {
    'advisory and extension':                               'Advisory and extension',
    'advisory & extension':                                 'Advisory and extension',
    'markets & finance':                                    'Markets & finance',
    'markets and finance':                                  'Markets & finance',
    'public health':                                        'Public health',
    'rural/public infrastructure':                          'Rural/public infrastructure',
    'rural/ public infrastructure':                         'Rural/public infrastructure',
    'behavior communication change & social programs':      'Behavior communication change & social programs',
    'food processing & storage':                            'Food processing & storage',
    'food processing and storage':                          'Food processing & storage',
    'livestock management & technologies':                  'Livestock Management & Technologies',
    'ecosystem services':                                   'Ecosystem services',
    'policy & regulatory':                                  'Policy & Regulatory',
    'crop management, technology, r&d':                     'Crop Management, Technology, R&D',
    'forestry management & technologies':                   'Forestry Management & Technologies',
    'aquaculture management & technologies':                'Aquaculture Management & Technologies',
}

OUT_ALIASES = {
    # curly apostrophe variant (common in copy-pasted text from Word/PDFs)
    'inclusivity & women\u2019s empowerment':               "Inclusivity & Women's Empowerment",
    # straight apostrophe variant
    "inclusivity & women's empowerment":                    "Inclusivity & Women's Empowerment",
    'sustainable economic growth':                          'Sustainable Economic Growth',
    'resilience':                                           'Resilience',
    'environmental sustainability':                         'Environmental Sustainability',
    'food security & nutrition':                            'Food Security & Nutrition',
    'food security and nutrition':                          'Food Security & Nutrition',
    'inducing policy change':                               'Inducing Policy Change',
}


# ──────────────────────────────────────────────
# 2. HELPER FUNCTIONS
# ──────────────────────────────────────────────

def normalize(s):
    """Lowercase and strip whitespace for consistent matching."""
    return s.strip().lower()


def map_cats(raw_value, aliases):
    """
    Split a potentially semicolon-separated category string,
    normalize each part, and return matching canonical names.
    Unrecognized values are silently skipped.
    """
    if pd.isna(raw_value):
        return []
    parts = [p.strip() for p in str(raw_value).split(';')]
    return [aliases[normalize(p)] for p in parts if normalize(p) in aliases]


def build_pivot(df, item_type, categories, aliases, all_projects):
    """
    Build a binary pivot table.
      - rows:    project IDs
      - columns: canonical category names
      - values:  1 if the project has that category, 0 otherwise
    """
    filtered = df[df['item_type'] == item_type].copy()
    pivot = pd.DataFrame(0, index=all_projects, columns=categories)
    pivot.index.name = 'project_id'

    for _, row in filtered.iterrows():
        pid = row['project_id']
        if pd.isna(pid):
            continue
        for cat in map_cats(row['primary_category'], aliases):
            pivot.loc[pid, cat] = 1
    # Add Total row
    pivot.loc['TOTAL'] = pivot.sum()
    return pivot


def build_cooccurrence(int_pivot, out_pivot, intervention_cats, outcome_cats):
    """
    Build a co-occurrence table counting how many projects
    have BOTH a given intervention AND a given outcome.
    """
    cooc = pd.DataFrame(0, index=intervention_cats, columns=outcome_cats)
    cooc.index.name = 'Intervention \\ Outcome'

    for pid in int_pivot.index:
        for i_cat in intervention_cats:
            if int_pivot.loc[pid, i_cat] == 1:
                for o_cat in outcome_cats:
                    if out_pivot.loc[pid, o_cat] == 1:
                        cooc.loc[i_cat, o_cat] += 1

    return cooc


# ──────────────────────────────────────────────
# 3. EXCEL STYLING HELPERS
# ──────────────────────────────────────────────

HEADER_FILL  = PatternFill('solid', start_color='1F4E79')
HEADER_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
INDEX_FILL   = PatternFill('solid', start_color='D6E4F0')
INDEX_FONT   = Font(bold=True, name='Arial', size=10)
ONE_FILL     = PatternFill('solid', start_color='C6EFCE')   # green for 1
ZERO_FILL    = PatternFill('solid', start_color='FFFFFF')   # white for 0
CELL_FONT    = Font(name='Arial', size=10)
CENTER       = Alignment(horizontal='center', vertical='center')
LEFT_WRAP    = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin         = Side(style='thin', color='BFBFBF')
BORDER       = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(cell, text):
    cell.value     = text
    cell.font      = HEADER_FONT
    cell.fill      = HEADER_FILL
    cell.alignment = CENTER
    cell.border    = BORDER


def write_pivot_sheet(ws, pivot_df, sheet_title):
    """Write a binary pivot DataFrame to an openpyxl worksheet."""
    ws.title = sheet_title
    cols = list(pivot_df.columns)

    # Header row
    style_header(ws.cell(1, 1), 'project_id')
    for ci, col in enumerate(cols, 2):
        style_header(ws.cell(1, ci), col)

    # Column widths
    ws.column_dimensions['A'].width = 14
    for ci in range(2, len(cols) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 22
    ws.row_dimensions[1].height = 50

    # Data rows
    for ri, (pid, row) in enumerate(pivot_df.iterrows(), 2):
        c = ws.cell(ri, 1, pid)
        c.font = INDEX_FONT; c.fill = INDEX_FILL
        c.alignment = CENTER; c.border = BORDER

        for ci, val in enumerate(row, 2):
            cell = ws.cell(ri, ci, int(val))
            cell.font      = CELL_FONT
            cell.fill      = ONE_FILL if val == 1 else ZERO_FILL
            cell.alignment = CENTER
            cell.border    = BORDER


def write_cooc_sheet(ws, cooc_df):
    """Write the co-occurrence table with a blue intensity gradient."""
    ws.title = 'Co-occurrence Table'
    cols = list(cooc_df.columns)
    max_val = cooc_df.values.max()

    style_header(ws.cell(1, 1), 'Intervention \\ Outcome')
    for ci, col in enumerate(cols, 2):
        style_header(ws.cell(1, ci), col)

    ws.column_dimensions['A'].width = 38
    for ci in range(2, len(cols) + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 24
    ws.row_dimensions[1].height = 50

    for ri, (icat, row) in enumerate(cooc_df.iterrows(), 2):
        c = ws.cell(ri, 1, icat)
        c.font = INDEX_FONT; c.fill = INDEX_FILL
        c.alignment = LEFT_WRAP; c.border = BORDER

        for ci, val in enumerate(row, 2):
            cell = ws.cell(ri, ci, int(val))
            cell.font = CELL_FONT
            cell.alignment = CENTER
            cell.border = BORDER

            # Blue gradient: low count = light blue, high count = dark blue
            intensity = val / max_val if max_val > 0 else 0
            r = int(255 - intensity * (255 - 30))
            g = int(255 - intensity * (255 - 80))
            b = int(255 - intensity * (255 - 162))
            cell.fill = PatternFill('solid', start_color=f'{r:02X}{g:02X}{b:02X}')


# ──────────────────────────────────────────────
# 4. HEATMAP
# ──────────────────────────────────────────────

def save_heatmap(cooc_df, output_path):
    """Save a matplotlib heatmap of the co-occurrence table."""
    fig, ax = plt.subplots(figsize=(13, 8))
    data = cooc_df.values.astype(float)

    im = ax.imshow(data, cmap='YlOrRd', aspect='auto')
    plt.colorbar(im, ax=ax, label='Co-occurrence count (# projects)')

    ax.set_xticks(range(len(cooc_df.columns)))
    ax.set_yticks(range(len(cooc_df.index)))
    ax.set_xticklabels(cooc_df.columns, rotation=30, ha='right', fontsize=10)
    ax.set_yticklabels(cooc_df.index, fontsize=10)
    ax.set_xlabel('Outcome Categories', fontsize=12, labelpad=10)
    ax.set_ylabel('Intervention Categories', fontsize=12, labelpad=10)
    ax.set_title('Intervention–Outcome Co-occurrence Heatmap',
                 fontsize=14, fontweight='bold', pad=15)

    # Annotate each cell with its count
    for i in range(len(cooc_df.index)):
        for j in range(len(cooc_df.columns)):
            val = int(data[i, j])
            color = 'white' if val > data.max() * 0.6 else 'black'
            ax.text(j, i, str(val), ha='center', va='center',
                    fontsize=9, color=color, fontweight='bold')

    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches='tight')
    plt.close()
    print(f"  Saved: {output_path}")


# ──────────────────────────────────────────────
# 5. MAIN
# ──────────────────────────────────────────────

def main():
    if not os.path.exists(INPUT_CSV_PATH):
        print(f"Error: file not found — {INPUT_CSV_PATH}")
        print("Please update INPUT_CSV_PATH at the top of this script.")
        sys.exit(1)

    output_dir   = os.path.dirname(os.path.abspath(INPUT_CSV_PATH))
    excel_path   = os.path.join(output_dir, 'livestock_pivot_tables.xlsx')
    heatmap_path = os.path.join(output_dir, 'intervention_outcome_heatmap.png')

    # ── Load data ──
    # header=2 skips the two blank rows at the top of this particular CSV
    print("Reading CSV...")
    df = pd.read_csv(INPUT_CSV_PATH, header=2)
    all_projects = sorted(df['project_id'].dropna().unique())
    print(f"  {len(all_projects)} unique projects, {len(df)} total rows")

    # ── Build pivots ──
    print("Building pivot tables...")
    int_pivot = build_pivot(df, 'intervention', INTERVENTION_CATS, INT_ALIASES, all_projects)
    out_pivot = build_pivot(df, 'outcome',       OUTCOME_CATS,      OUT_ALIASES, all_projects)
    cooc      = build_cooccurrence(int_pivot, out_pivot, INTERVENTION_CATS, OUTCOME_CATS)

    print(f"  Interventions per category:\n{int_pivot.sum().to_string()}")
    print(f"  Outcomes per category:\n{out_pivot.sum().to_string()}")

    # ── Write three pivot tables to a single Excel Workbook ──
    print("Writing Excel file...")
    wb = openpyxl.Workbook()

    ws1 = wb.active
    write_pivot_sheet(ws1, int_pivot, 'Interventions Pivot')

    ws2 = wb.create_sheet()
    write_pivot_sheet(ws2, out_pivot, 'Outcomes Pivot')

    ws3 = wb.create_sheet()
    write_cooc_sheet(ws3, cooc)

    ## Individual CSVs (optional, can be commented out if not needed)
    int_pivot.to_csv(os.path.join(output_dir, 'interventions_pivot.csv'))
    out_pivot.to_csv(os.path.join(output_dir, 'outcomes_pivot.csv'))
    cooc.to_csv(os.path.join(output_dir, 'cooccurrence_table.csv'))

    wb.save(excel_path)
    print(f"  Saved: {excel_path}")

    # ── Save heatmap ──
    print("Generating heatmap...")
    save_heatmap(cooc, heatmap_path)

    print("\nDone! Outputs saved to:", output_dir)


if __name__ == '__main__':
    main()